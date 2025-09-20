import os, yaml, time
import numpy as np
from pathlib import Path
from baselines.EventBaselineLab import EventBaseline
from baselines.download_baseline import clone_repo
from baselines.VPR_Tutorial.evaluation.metrics import recallAtK, createPR
import prettytable
import openpyxl
from datetime import datetime
import re

class name_baseline(EventBaseline):
    def __init__(self):
        super().__init__()

        self.name = "<name>"
        # Check if the baseline repository is already cloned
        self.repo_path = "./baselines/<name>"
        # Baseline URL
        self.url = "https://github.com/<name>"
        if not os.path.exists(self.repo_path):
            clone_repo(self.url, destination=self.repo_path)
        self.baseline_config_path = './baselines/<name>.yaml'
        # Load the baseline configuration
        with open(self.baseline_config_path, 'r') as file:
            self.baseline_config = yaml.safe_load(file)
        # Create the data output folder
        self.outdir = './output/<name>'
        os.makedirs(self.outdir, exist_ok=True)

    def format_data(self, config, reference, query, timewindow):
        """
        Format the reference and query data for the baseline.
        """
        # Get experimental details
        ref_info = reference.get_dataset_info()
        query_info = query.get_dataset_info()

        ref_name = ref_info['sequence_name']
        query_name = query_info['sequence_name']

        # from ref_info['file_path'] dict, find the directory that matches ref/query name and timewindow
        self.ref_key = [d for d in ref_info['file_path'] if ref_name in d and str(timewindow) in d]
        self.query_key = [d for d in query_info['file_path'] if query_name in d and str(timewindow) in d]
        self.ref_directory = ref_info['file_path'][self.ref_key[0]]
        self.query_directory = query_info['file_path'][self.query_key[0]]

        # Get the reference and query data
        ref_files = sorted(list(Path(self.ref_directory).glob("*.npy")))
        query_files = sorted(list(Path(self.query_directory).glob("*.npy")))

        # Load the reference and query data as numpy arrays
        self.reference_data = np.array([np.load(ref_file) for ref_file in ref_files])
        self.query_data = np.array([np.load(query_file) for query_file in query_files])

        # Set the output folder
        self.output_dir = os.path.join(self.outdir, f"{ref_info['dataset_name']}", f"{ref_info['sequence_name']}_{query_info['sequence_name']}",
                                       f"{config['frame_generator']}_{timewindow}")
        os.makedirs(self.output_dir, exist_ok=True)

    def build_execute(self, config, data_config, ground_truth):
        """
        Build a commandline execute for the baseline with the provided reference, query, and ground truth data.
        """
        pass

    def run(self):
        """
        Run the baseline.
        """
        '''
        Implement run logic here to retrieve distance matrix and save it for analysis.
        '''
        # Save the distance matrices
        np.save(f"{self.output_dir}/distance_matrix.npy", distance_matrix)
    
    def parse_results(self, GT):
        """
        Summary sheet: upsert by (run_name, ref_query, array_name)
        Per-run sheet (self.name): upsert summary by (ref_query, array_name),
        and upsert each PR block keyed by "PR curve for {ref_query} :: {array_name}".
        """
        # gather files
        all_files = sorted(list(Path(self.output_dir).glob("*.npy")))
        all_names = [os.path.basename(f).replace(".npy", "") for f in all_files]
        all_arrays = [np.load(f) for f in all_files]
        GThard = np.load(GT)
        if not all_arrays:
            print("No .npy result files found in", self.output_dir)
            return

        # conform GT shape
        target_shape = all_arrays[0].shape
        if GThard.shape != target_shape:
            slices = tuple(slice(0, min(GThard.shape[i], target_shape[i])) for i in range(len(target_shape)))
            GThard = GThard[slices]

        K_list = [1, 5, 10, 15, 20, 25]
        try:
            table = prettytable.PrettyTable()
            table.field_names = ["Recall@K"] + [f"@{k}" for k in K_list] + ["AUPR"]
        except Exception:
            table = None

        run_name = getattr(self, "name", None) or f"{getattr(self, 'ref_name', '')}__{getattr(self, 'query_name', '')}"
        ref_query = f"{getattr(self, 'ref_name', '')}__{getattr(self, 'query_name', '')}"
        timestamp = datetime.utcnow().isoformat()

        rows = []
        pr_curves = {}  # (ref_query, array_name) -> (P,R)

        for name, array in zip(all_names, all_arrays):
            recalls = []
            for k in K_list:
                try:
                    r = recallAtK(array, GThard, K=k)
                except Exception as e:
                    print(f"  -> Error computing recallAtK for {name} K={k}: {e}")
                    r = np.nan
                recalls.append(np.round(r, 2))

            try:
                P, R = createPR(array, GThard, matching='single', n_thresh=100)
                P = np.asarray(P); R = np.asarray(R)
                idx = np.argsort(R)
                aupr = float(np.trapz(P[idx], R[idx]))
            except Exception as e:
                print(f"  -> Error computing PR for {name}: {e}")
                P, R, aupr = np.array([]), np.array([]), np.nan

            if table is not None:
                table.add_row([name] + recalls + [np.round(aupr, 4)])

            rows.append({
                "timestamp_utc": timestamp,
                "run_name": run_name,
                "ref_query": ref_query,
                "array_name": name,
                "n_references": int(target_shape[0]) if len(target_shape) >= 1 else None,
                "n_queries": int(target_shape[1]) if len(target_shape) >= 2 else None,
                "R@1": recalls[0], "R@5": recalls[1], "R@10": recalls[2],
                "R@15": recalls[3], "R@20": recalls[4], "R@25": recalls[5],
                "aupr": np.round(aupr, 6)
            })
            pr_curves[(ref_query, name)] = (P, R)

        if table is not None:
            print(table)

        # --- Excel I/O (openpyxl only; no sheet deletions) ---
        excel_path = "./output/eventlab_results.xlsx"
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)

        def load_wb(path):
            try:
                if os.path.exists(path):
                    return openpyxl.load_workbook(path)
            except Exception:
                bak = f"{path}.corrupt.{int(time.time())}.bak"
                os.rename(path, bak)
            return openpyxl.Workbook()

        def sheet_headers(ws, headers):
            existing = [ws.cell(row=1, column=i+1).value for i in range(len(headers))]
            if existing != headers:
                for i, h in enumerate(headers, 1):
                    ws.cell(row=1, column=i, value=h)
            return {h: i+1 for i, h in enumerate(headers)}

        def build_index(ws, key_cols):  # returns {key_tuple: row_idx}
            idx = {}
            r = 2
            while r <= ws.max_row:
                vals = [ws.cell(row=r, column=c).value for c in key_cols]
                if all(v in (None, "") for v in vals):
                    # stop only if an entirely empty line AND next line empty too
                    nxt = [ws.cell(row=r+1, column=c).value for c in key_cols] if r+1 <= ws.max_row else [None]
                    if all(v in (None, "") for v in nxt):
                        break
                else:
                    idx[tuple(vals)] = r
                r += 1
            return idx

        def upsert_rows(ws, headers, rows_to_write, match_keys):
            hdr = sheet_headers(ws, headers)
            key_cols = [hdr[k] for k in match_keys]
            existing = build_index(ws, key_cols)

            # how many new rows will be appended?
            new_count = sum(1 for d in rows_to_write if tuple(d[k] for k in match_keys) not in existing)

            # if there is content immediately after current data, insert space for new rows
            if new_count > 0:
                last_existing_row = max(existing.values()) if existing else 1
                tail_row = last_existing_row + 1
                has_content_below = any(ws.cell(row=tail_row, column=c).value not in (None, "")
                                        for c in range(1, ws.max_column + 1))
                if has_content_below:
                    ws.insert_rows(tail_row, amount=new_count)
                    existing = build_index(ws, key_cols)  # rebuild after shift

            next_row = (max(existing.values()) + 1) if existing else 2
            for d in rows_to_write:
                key = tuple(d[k] for k in match_keys)
                r = existing.get(key)
                if r is None:
                    r = next_row
                    next_row += 1
                for h in headers:
                    ws.cell(row=r, column=hdr[h], value=d.get(h, None))

        def find_text_row(ws, text):
            for r in range(1, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == text:
                    return r
            return None

        def write_pr_table(ws, start_row, P, R):
            ws.cell(row=start_row, column=1, value="precision")
            ws.cell(row=start_row, column=2, value="recall")
            for i in range(len(P)):
                ws.cell(row=start_row + 1 + i, column=1, value=float(P[i]))
                ws.cell(row=start_row + 1 + i, column=2, value=float(R[i]))

        def upsert_pr_block(ws, header_text, P, R):
            # create/update a PR block under its own header line
            header_row = find_text_row(ws, header_text)
            if header_row is None:
                start = ws.max_row + (0 if (ws.max_row == 1 and ws.cell(1,1).value in (None,"")) else 2)
                ws.cell(row=start, column=1, value=header_text)
                write_pr_table(ws, start + 1, P, R)
                return

            data_start = header_row + 2
            r = data_start
            while r <= ws.max_row:
                val = ws.cell(row=r, column=1).value
                if isinstance(val, str) and val.startswith("PR curve for "):
                    break
                if val in (None, "") and ws.cell(row=r+1, column=1).value in (None, ""):
                    break
                r += 1
            old_len = max(0, r - data_start)
            new_len = len(P)

            if new_len > old_len:
                ws.insert_rows(data_start, amount=(new_len - old_len))
            elif new_len < old_len:
                ws.delete_rows(data_start, amount=(old_len - new_len))

            write_pr_table(ws, header_row + 1, P, R)

        wb = load_wb(excel_path)

        # Summary (global)
        summary_headers = ["timestamp_utc","run_name","ref_query","array_name","n_references","n_queries",
                        "R@1","R@5","R@10","R@15","R@20","R@25","aupr"]
        ws_sum = wb["Summary"] if "Summary" in wb.sheetnames else wb.create_sheet("Summary")
        sheet_headers(ws_sum, summary_headers)
        upsert_rows(ws_sum, summary_headers, rows, match_keys=["run_name","ref_query","array_name"])

        # Per-run sheet
        safe_run = re.sub(r'[:\\/?*\[\]]', "_", str(run_name)).strip()[:31] or "run"
        ws_run = wb[safe_run] if safe_run in wb.sheetnames else wb.create_sheet(safe_run)
        run_headers = summary_headers[:]  # same columns
        sheet_headers(ws_run, run_headers)
        upsert_rows(ws_run, run_headers, rows, match_keys=["ref_query","array_name"])

        # PR blocks: unique per (ref_query, array_name)
        for (rq, arr), (P, R) in pr_curves.items():
            header_text = f"PR curve for {rq} :: {arr}"
            upsert_pr_block(ws_run, header_text, P, R)

        # tidy default empty sheet if present
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1 and wb["Sheet"].cell(1,1).value in (None,""):
            del wb["Sheet"]

        wb.save(excel_path)
        print(f"Saved Excel workbook to: {excel_path}")

    def cleanup(self):
        """
        Clean up temporary files.
        """
        import shutil
        if hasattr(self, 'temp_dir') and os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
import prettytable, os

import numpy as np
from skimage.transform import resize

from baselines.VPR_Tutorial.evaluation.metrics import recallAtK, createPR
from utils.functional import create_GTtol
import matplotlib.pyplot as plt
from matplotlib.patches import Patch

def overlay_matches_on_array(
    array,
    GThard,
    top_k=1,
    pred_mode="per_column",   # "per_column" or "per_row"
    marker_size=20,
    alpha_blend=0.7,
    save_path=None,
    array_name=None,
):
    """
    Create an RGB overlay of the similarity/distance matrix with TP/FP/FN markers.

    - array: 2D numpy array (refs x queries)
    - GThard: same-shaped 0/1 GT array (or will be resized)
    - top_k: how many predictions per unit (row or column depending on pred_mode)
    - pred_mode: 'per_column' (default) or 'per_row'
        * per_column: for each query column, select top_k reference rows (ensures predictions are chosen per query)
        * per_row: for each reference row, select top_k query cols (previous behaviour)
    - matrix_type: "distance" or "similarity"
    """
    if array.ndim != 2:
        raise ValueError("array must be 2D (refs x queries)")

    # Resize GT if needed (same behaviour as your run_metrics)
    if GThard.shape != array.shape:
        GT = resize(GThard, array.shape, order=0, preserve_range=True, anti_aliasing=False)
        GT = (GT > 0.5).astype(int)
    else:
        GT = (GThard > 0.5).astype(int)

    h, w = array.shape

    # Build predictions according to pred_mode
    preds_idx = []
    top_k = max(1, int(top_k))
    if pred_mode == "per_column":
        # For each column, find the best rows
        # argsort descending along axis=0 (rows sorted per column)
        sorted_rows_per_col = np.argsort(-array, axis=0)  # shape (h, w)
        for c in range(w):
            for k in range(min(top_k, h)):
                r = sorted_rows_per_col[k, c]
                preds_idx.append((r, c))
    elif pred_mode == "per_row":
        # For each row, find best columns (old behaviour)
        sorted_cols_per_row = np.argsort(-array, axis=1)  # shape (h, w)
        for r in range(h):
            for k in range(min(top_k, w)):
                c = sorted_cols_per_row[r, k]
                preds_idx.append((r, c))
    else:
        raise ValueError("pred_mode must be 'per_column' or 'per_row'")

    # Deduplicate predictions (just in case) and keep ordering
    seen = set()
    preds_unique = []
    for p in preds_idx:
        if p not in seen:
            preds_unique.append(p)
            seen.add(p)
    preds_idx = preds_unique

    # Compute TP, FP, FN
    tp_idx = [(r, c) for (r, c) in preds_idx if GT[r, c] == 1]
    fp_idx = [(r, c) for (r, c) in preds_idx if GT[r, c] == 0]

    gt_pairs = list(zip(*np.nonzero(GT)))  # all GT==1 pairs
    predicted_set = set(preds_idx)
    fn_idx = [pair for pair in gt_pairs if pair not in predicted_set]

    # Build grayscale RGB base
    base = np.stack([array, array, array], axis=2)
    overlay = base.copy()

    # Blend small color contribution into overlay at predicted/gt locations so shading remains visible
    def blend_color(positions, color_rgb):
        for (r, c) in positions:
            if 0 <= r < h and 0 <= c < w:
                overlay[r, c, :] = (1.0 - alpha_blend) * overlay[r, c, :] + alpha_blend * np.array(color_rgb)

    GREEN = (0.0, 1.0, 0.0)
    RED   = (1.0, 0.0, 0.0)
    BLUE  = (0.0, 0.4, 1.0)

    blend_color(tp_idx, GREEN)
    blend_color(fp_idx, RED)
    blend_color(fn_idx, BLUE)

    overlay_rgb = (np.clip(overlay, 0.0, 1.0) * 255).astype(np.uint8)

    fig, ax = plt.subplots(figsize=(8, 6))
    ax.imshow(overlay_rgb, origin="upper", interpolation="nearest")
    ax.set_title(f"Matches overlay (mode={pred_mode}, top_k={top_k})")
    ax.set_xlabel("Query index (cols)")
    ax.set_ylabel("Reference index (rows)")

    # plot outline markers so they are visible on top of blended pixels
    if tp_idx:
        ys, xs = zip(*tp_idx)
        ax.scatter(xs, ys, s=marker_size, facecolors='none', edgecolors='lime', linewidths=0.9, label='TP')
    if fp_idx:
        ys, xs = zip(*fp_idx)
        ax.scatter(xs, ys, s=marker_size, facecolors='none', edgecolors='red', linewidths=0.9, label='FP')

    legend_handles = [
        Patch(edgecolor='lime', facecolor='none', label=f'TP ({len(tp_idx)})'),
        Patch(edgecolor='red',  facecolor='none', label=f'FP ({len(fp_idx)})'),
    ]
    ax.legend(handles=legend_handles, loc='upper right', framealpha=0.9)
    plt.tight_layout()

    if save_path:
        fig.savefig(os.path.join(f'{save_path}', f'{array_name}_matches') , dpi=200)

    return fig, preds_idx, tp_idx, fp_idx, fn_idx

class EventBaseline:
    def __init__(self):
        self.K_list = [1, 5, 10, 15, 20, 25]

    def run_metrics(self, all_names, all_arrays, GThard, timestamp, run_name, ref_query, matrix_type="distance", outdir=None, tolerance=0):
        # Create a pretty table for displaying results
        table = prettytable.PrettyTable()
        table.field_names = ["Recall@K"] + [f"@{k}" for k in self.K_list] + ["AUPR"]
        rows = []
        pr_curves = {}  # (ref_query, array_name) -> (P,R)
        GT = None
        # if the name is `all`, use the GThard_noseq instead
        for name, array in zip(all_names, all_arrays):
            recalls = []
            if matrix_type == "distance":
                array = array.max() - array  # convert to similarity
            for k in self.K_list:
                # Ensure the GThard shape matches the array shape
                target_shape = array.shape
                if GThard.shape != target_shape:
                    # use numpy reshape
                    GT = resize(GThard, target_shape, order=0,
                                preserve_range=True, anti_aliasing=False)
                    # Apply ground truth tolerance
                    GT = (GT > 0.5).astype(int)
                    GT_tol = create_GTtol(GT, tolerance)
                else:
                    GT_tol = create_GTtol(GThard, tolerance)
                r = recallAtK(array, GT_tol, K=k)
                recalls.append(np.round(r, 2))
            
            overlay_rgb, preds, tp, fp, fn = overlay_matches_on_array(
                array=array,
                GThard=GT_tol,
                top_k=1,                 # set to 1 to get single predicted ref per query
                pred_mode="per_column",  # important: choose 'per_column' to select predictions per query
                marker_size=20,
                alpha_blend=0.6,
                save_path=outdir,
                array_name=name
            )

            try:
                P, R = createPR(array, GT_tol, matching='single', n_thresh=100)
                P = np.asarray(P); R = np.asarray(R)
                idx = np.argsort(R)
                aupr = float(np.trapz(P[idx], R[idx]))
            except Exception as e:
                print(f"  -> Error computing PR for {name}: {e}")
                P, R, aupr = np.array([]), np.array([]), np.nan

            if table is not None:
                table.add_row([name] + recalls + [np.round(aupr, 4)])

            rows.append({
                "timestamp_utc": timestamp,
                "run_name": run_name,
                "ref_query": ref_query,
                "array_name": name,
                "n_references": int(target_shape[0]) if len(target_shape) >= 1 else None,
                "n_queries": int(target_shape[1]) if len(target_shape) >= 2 else None,
                "R@1": recalls[0], "R@5": recalls[1], "R@10": recalls[2],
                "R@15": recalls[3], "R@20": recalls[4], "R@25": recalls[5],
                "aupr": np.round(aupr, 6)
            })
            pr_curves[(ref_query, name)] = (P, R)

        if table is not None:
            print(table)
        
        return rows, pr_curves
    
    def save_results(self, rows, pr_curves, run_name, ref_query):
        """
        rows: list[dict] with keys:
            timestamp_utc, run_name, ref_query, array_name, n_references, n_queries,
            R@1,R@5,R@10,R@15,R@20,R@25, aupr
        pr_curves: dict[(ref_query, array_name)] -> (P, R)
        """
        import os, re, time
        from datetime import datetime, timezone
        import openpyxl

        # ---------- paths ----------
        excel_path = "./output/eventlab_results.xlsx"
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)

        # ---------- workbook/sheet helpers ----------
        def load_wb(path):
            try:
                if os.path.exists(path):
                    wb_ = openpyxl.load_workbook(path)
                else:
                    wb_ = openpyxl.Workbook()
                # remove empty default "Sheet" if present
                if "Sheet" in wb_.sheetnames:
                    ws0 = wb_["Sheet"]
                    if ws0.max_row == 1 and ws0.max_column == 1 and ws0.cell(1,1).value in (None, ""):
                        del wb_["Sheet"]
                return wb_
            except Exception:
                bak = f"{path}.corrupt.{int(time.time())}.bak"
                try:
                    os.rename(path, bak)
                except Exception:
                    pass
                wb_ = openpyxl.Workbook()
                if "Sheet" in wb_.sheetnames:
                    del wb_["Sheet"]
                return wb_

        def ensure_sheet(wb, name):
            return wb[name] if name in wb.sheetnames else wb.create_sheet(name)

        def sheet_headers(ws, headers):
            existing = [ws.cell(row=1, column=i+1).value for i in range(len(headers))]
            if existing != headers:
                for i, h in enumerate(headers, 1):
                    ws.cell(row=1, column=i, value=h)
            return {h: i+1 for i, h in enumerate(headers)}

        # --- index helper used by Summary only (run sheet uses append-only upsert) ---
        def build_index(ws, key_cols):  # returns {key_tuple: row_idx}
            idx = {}
            r = 2
            while r <= ws.max_row:
                vals = [ws.cell(row=r, column=c).value for c in key_cols]
                if all(v in (None, "") for v in vals):
                    nxt = [ws.cell(row=r+1, column=c).value for c in key_cols] if r+1 <= ws.max_row else [None]
                    if all(v in (None, "") for v in nxt):
                        break
                else:
                    idx[tuple(vals)] = r
                r += 1
            return idx

        # ---------- Per-run sheet: append-only upsert (NO ROW INSERTIONS) ----------
        # --- new helper: ALWAYS APPEND (no matching/overwrite) ---
        def append_rows_no_dedupe(ws, headers, rows_to_write):
            hdr = sheet_headers(ws, headers)
            next_row = ws.max_row + 1 if ws.max_row >= 1 else 2
            for d in rows_to_write:
                for h in headers:
                    ws.cell(row=next_row, column=hdr[h], value=d.get(h, None))
                next_row += 1


        # ---------- PR columns on the SAME run sheet (no row sharing issues) ----------
        def pr_find_block_col(ws, run_headers, run_label, array_name):
            """Find starting column of an existing PR block on the run sheet."""
            hp = f"PR ({run_label}) [{array_name}] - Precision"
            hr = f"PR ({run_label}) [{array_name}] - Recall"
            for col in range(len(run_headers)+2, ws.max_column):  # search to the right of recall table
                if ws.cell(row=1, column=col).value == hp and ws.cell(row=1, column=col+1).value == hr:
                    return col
            return None

        def pr_next_free_col(ws, start_col):
            """Return first empty *pair* of columns at/after start_col (row 1 empty in both)."""
            col = max(1, start_col)
            while True:
                c1 = ws.cell(row=1, column=col).value
                c2 = ws.cell(row=1, column=col+1).value
                if (c1 in (None, "")) and (c2 in (None, "")):
                    return col
                col += 2

        def pr_clear_block(ws, start_col, top_row=2, n_rows=5000):
            """Clear a tall region under the headers (no row shifts)."""
            for r in range(top_row, top_row + n_rows):
                ws.cell(row=r, column=start_col,     value=None)
                ws.cell(row=r, column=start_col + 1, value=None)

        def pr_write_block_on_run_sheet(ws, run_headers, run_label, array_name, P, R):
            """
            Write/overwrite a 2-col PR block for (run_label, array_name) on the SAME run sheet.
            Headers at row 1; numeric data from row 2 down. No gaps, no inserts.
            """
            base_cols = len(run_headers)  # PR always starts to the right of these
            start_col = pr_find_block_col(ws, run_headers, run_label, array_name)
            if start_col is None:
                start_col = pr_next_free_col(ws, base_cols + 2)

            # headers on row 1
            ws.cell(row=1, column=start_col,     value=f"PR ({run_label}) [{array_name}] - Precision")
            ws.cell(row=1, column=start_col + 1, value=f"PR ({run_label}) [{array_name}] - Recall")

            # clear old contents below headers
            pr_clear_block(ws, start_col, top_row=2, n_rows=5000)

            # write numeric data (row 2..)
            n = max(len(P), len(R))
            for i in range(n):
                ws.cell(row=2 + i, column=start_col,     value=float(P[i]) if i < len(P) else None)
                ws.cell(row=2 + i, column=start_col + 1, value=float(R[i]) if i < len(R) else None)

        def pr_write_all_blocks_on_run_sheet(ws_run, pr_curves, run_headers, run_label):
            # deterministic by array_name
            items = sorted(pr_curves.items(), key=lambda x: x[0][1])  # ((ref_query, array_name), (P,R))
            for (rq, array_name), (P, R) in items:
                pr_write_block_on_run_sheet(ws_run, run_headers, run_label, array_name, P, R)

        # ---------- Aggregated Summary (across timewindows) ----------
        METRICS = ["R@1","R@5","R@10","R@15","R@20","R@25","aupr"]

        summary_headers = (
            ["timestamp_utc","run_name","ref_query","array_name","n_references","n_queries"] +
            [f"{m}_mean" for m in METRICS] +
            [f"{m}_std"  for m in METRICS] +
            [f"{m}_n"    for m in METRICS]
        )

        def clean_label(label: str) -> str:
            """Remove -frames-### / -reconstruction-### so timewindows aggregate."""
            if label is None:
                return ""
            return re.sub(r"-(?:frames|reconstruction)-\d+", "", str(label))

        def update_agg_cell(ws, row_idx, hdr_map, base, new_val):
            """Welford (sample std) incremental update."""
            if new_val is None:
                return
            c_mean, c_std, c_n = hdr_map[f"{base}_mean"], hdr_map[f"{base}_std"], hdr_map[f"{base}_n"]

            try: old_n = int(ws.cell(row=row_idx, column=c_n).value or 0)
            except Exception: old_n = 0
            try: old_mean = float(ws.cell(row=row_idx, column=c_mean).value)
            except Exception: old_mean = None
            try: old_std = float(ws.cell(row=row_idx, column=c_std).value)
            except Exception: old_std = None

            if old_n < 1 or old_mean is None or old_std is None:
                ws.cell(row=row_idx, column=c_n,    value=1)
                ws.cell(row=row_idx, column=c_mean, value=float(new_val))
                ws.cell(row=row_idx, column=c_std,  value=0.0)
                return

            M2 = (old_std ** 2) * (old_n - 1) if old_n > 1 else 0.0
            n_new = old_n + 1
            delta = float(new_val) - old_mean
            mean_new = old_mean + delta / n_new
            M2_new = M2 + delta * (float(new_val) - mean_new)
            std_new = (M2_new / (n_new - 1)) ** 0.5 if n_new > 1 else 0.0

            ws.cell(row=row_idx, column=c_n,    value=n_new)
            ws.cell(row=row_idx, column=c_mean, value=mean_new)
            ws.cell(row=row_idx, column=c_std,  value=std_new)

        def find_or_create_summary_row(ws_sum, hdr_map, clean_run, clean_rq, array_name, n_refs, n_qs):
            key_cols = [hdr_map["run_name"], hdr_map["ref_query"], hdr_map["array_name"]]
            existing = build_index(ws_sum, key_cols)
            key_tuple = (clean_run, clean_rq, array_name)
            r = existing.get(key_tuple)
            if r is None:
                r = (max(existing.values()) + 1) if existing else 2
                ws_sum.cell(row=r, column=hdr_map["timestamp_utc"], value=datetime.now(timezone.utc).isoformat())
                ws_sum.cell(row=r, column=hdr_map["run_name"],     value=clean_run)
                ws_sum.cell(row=r, column=hdr_map["ref_query"],    value=clean_rq)
                ws_sum.cell(row=r, column=hdr_map["array_name"],   value=array_name)
                ws_sum.cell(row=r, column=hdr_map["n_references"], value=n_refs)
                ws_sum.cell(row=r, column=hdr_map["n_queries"],    value=n_qs)
                for m in METRICS:
                    ws_sum.cell(row=r, column=hdr_map[f"{m}_n"], value=0)
            else:
                ws_sum.cell(row=r, column=hdr_map["timestamp_utc"], value=datetime.now(timezone.utc).isoformat())
                ws_sum.cell(row=r, column=hdr_map["n_references"], value=n_refs)
                ws_sum.cell(row=r, column=hdr_map["n_queries"],    value=n_qs)
            return r
        
        def append_rows_force(ws, headers, rows_to_write):
            hdr = sheet_headers(ws, headers)

            # Find the last *actually used* row across the first N columns
            def last_used_row(ws, ncols):
                r = ws.max_row or 1
                # scan upward until we hit a row that has any value in cols 1..ncols
                while r >= 2:
                    if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, ncols+1)):
                        return r
                    r -= 1
                return 1  # only header remains

            start_row = last_used_row(ws, len(headers)) + 1
            r = start_row
            for d in rows_to_write:
                for h in headers:
                    ws.cell(row=r, column=hdr[h], value=d.get(h, None))
                r += 1


        # ---------- write workbook ----------
        wb = load_wb(excel_path)

        # Summary sheet (aggregated)
        ws_sum = ensure_sheet(wb, "Summary")
        hdr_sum = sheet_headers(ws_sum, summary_headers)

        # Per-run sheet (recall rows + PR columns on the right)
        safe_run = re.sub(r'[:\\/?*\[\]]', "_", str(run_name)).strip()[:31] or "run"
        ws_run = ensure_sheet(wb, safe_run)

        run_headers = [
            "timestamp_utc","run_name","ref_query","array_name","n_references","n_queries",
            "R@1","R@5","R@10","R@15","R@20","R@25","aupr"
        ]
        sheet_headers(ws_run, run_headers)

        # 1) Upsert recall rows WITHOUT INSERTING ROWS (so PR columns never shift)
        append_rows_force(ws_run, run_headers, rows)

        # 2) PR columns to the RIGHT on the SAME sheet (unique per run/timewindow & array)
        run_label_for_block = f"{run_name} :: {ref_query}"
        pr_write_all_blocks_on_run_sheet(ws_run, pr_curves, run_headers, run_label_for_block)

        # 3) Incremental aggregation into Summary (across timewindows)
        for row in rows:
            clean_rq  = clean_label(row["ref_query"])
            clean_run = clean_label(row["run_name"])
            r_idx = find_or_create_summary_row(
                ws_sum, hdr_sum, clean_run, clean_rq, row["array_name"],
                n_refs=row.get("n_references"), n_qs=row.get("n_queries")
            )
            for m in METRICS:
                update_agg_cell(ws_sum, r_idx, hdr_sum, m, row.get(m))

        wb.save(excel_path)
        print(f"Saved Excel workbook to: {excel_path}")
